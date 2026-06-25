from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
import sqlite3, json, io, re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI(title="Pulim API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

DB = "pulim.db"

# ─── DATABASE SETUP ───────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            telegram_id TEXT PRIMARY KEY,
            name        TEXT,
            theme       TEXT DEFAULT 'dark',
            created_at  TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS budgets (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id TEXT,
            year        INTEGER,
            month       INTEGER,
            amount      REAL DEFAULT 0,
            UNIQUE(telegram_id, year, month)
        );
        CREATE TABLE IF NOT EXISTS transactions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id TEXT,
            type        TEXT,
            amount      REAL,
            name        TEXT,
            cat_id      TEXT,
            date        TEXT,
            created_at  TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS categories (
            id          TEXT PRIMARY KEY,
            telegram_id TEXT,
            cat_type    TEXT,
            emoji       TEXT,
            name        TEXT,
            position    INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS wishes (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id TEXT,
            emoji       TEXT,
            name        TEXT,
            price       REAL,
            created_at  TEXT DEFAULT (datetime('now'))
        );
    """)
    conn.commit()
    conn.close()

init_db()

# ─── MODELS ───────────────────────────────────────────────
class UserIn(BaseModel):
    telegram_id: str
    name: str
    theme: Optional[str] = "dark"

class BudgetIn(BaseModel):
    telegram_id: str
    year: int
    month: int
    amount: float

class TransactionIn(BaseModel):
    telegram_id: str
    type: str
    amount: float
    name: str
    cat_id: str
    date: str

class CategoryIn(BaseModel):
    telegram_id: str
    cat_type: str
    emoji: str
    name: str

class CategoryReorder(BaseModel):
    telegram_id: str
    cat_type: str
    ordered_ids: List[str]

class WishIn(BaseModel):
    telegram_id: str
    emoji: str
    name: str
    price: float

class ThemeIn(BaseModel):
    telegram_id: str
    theme: str

# ─── DEFAULT CATEGORIES ───────────────────────────────────
DEFAULT_EXPENSE_CATS = [
    ("🍔","Еда"), ("🚕","Транспорт"), ("🏠","Аренда"),
    ("💊","Здоровье"), ("🎬","Развлечения"), ("👗","Одежда"), ("📚","Обучение"),
]
DEFAULT_INCOME_CATS = [
    ("💼","Зарплата"), ("💻","Фриланс"), ("📈","Другое"),
]

def ensure_user_categories(telegram_id: str, conn):
    c = conn.cursor()
    existing = c.execute("SELECT id FROM categories WHERE telegram_id=?", (telegram_id,)).fetchall()
    if not existing:
        import uuid
        for i, (emoji, name) in enumerate(DEFAULT_EXPENSE_CATS):
            c.execute("INSERT INTO categories(id,telegram_id,cat_type,emoji,name,position) VALUES(?,?,?,?,?,?)",
                      (str(uuid.uuid4())[:8], telegram_id, "expense", emoji, name, i))
        for i, (emoji, name) in enumerate(DEFAULT_INCOME_CATS):
            c.execute("INSERT INTO categories(id,telegram_id,cat_type,emoji,name,position) VALUES(?,?,?,?,?,?)",
                      (str(uuid.uuid4())[:8], telegram_id, "income", emoji, name, i))
        conn.commit()

# ─── USERS ────────────────────────────────────────────────
@app.post("/users")
def upsert_user(u: UserIn):
    conn = get_db()
    conn.execute("""
        INSERT INTO users(telegram_id,name,theme) VALUES(?,?,?)
        ON CONFLICT(telegram_id) DO UPDATE SET name=excluded.name
    """, (u.telegram_id, u.name, u.theme))
    conn.commit()
    ensure_user_categories(u.telegram_id, conn)
    conn.close()
    return {"ok": True}

@app.get("/users/{telegram_id}")
def get_user(telegram_id: str):
    conn = get_db()
    ensure_user_categories(telegram_id, conn)
    row = conn.execute("SELECT * FROM users WHERE telegram_id=?", (telegram_id,)).fetchone()
    conn.close()
    if not row:
        return {"exists": False}
    return {"exists": True, **dict(row)}

@app.patch("/users/theme")
def update_theme(t: ThemeIn):
    conn = get_db()
    conn.execute("UPDATE users SET theme=? WHERE telegram_id=?", (t.theme, t.telegram_id))
    conn.commit()
    conn.close()
    return {"ok": True}

# ─── BUDGET ───────────────────────────────────────────────
@app.post("/budget")
def set_budget(b: BudgetIn):
    conn = get_db()
    conn.execute("""
        INSERT INTO budgets(telegram_id,year,month,amount) VALUES(?,?,?,?)
        ON CONFLICT(telegram_id,year,month) DO UPDATE SET amount=excluded.amount
    """, (b.telegram_id, b.year, b.month, b.amount))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.get("/budget")
def get_budget(telegram_id: str, year: int, month: int):
    conn = get_db()
    row = conn.execute("SELECT amount FROM budgets WHERE telegram_id=? AND year=? AND month=?",
                       (telegram_id, year, month)).fetchone()
    conn.close()
    return {"amount": row["amount"] if row else 0}

# ─── TRANSACTIONS ─────────────────────────────────────────
@app.post("/transactions")
def add_transaction(t: TransactionIn):
    conn = get_db()
    cur = conn.execute("""
        INSERT INTO transactions(telegram_id,type,amount,name,cat_id,date)
        VALUES(?,?,?,?,?,?)
    """, (t.telegram_id, t.type, t.amount, t.name, t.cat_id, t.date))
    conn.commit()
    tx_id = cur.lastrowid
    conn.close()
    return {"ok": True, "id": tx_id}

@app.get("/transactions")
def get_transactions(telegram_id: str, year: int, month: int):
    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM transactions
        WHERE telegram_id=? AND strftime('%Y',date)=? AND strftime('%m',date)=?
        ORDER BY date DESC, created_at DESC
    """, (telegram_id, str(year), f"{month:02d}")).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.delete("/transactions/{tx_id}")
def delete_transaction(tx_id: int, telegram_id: str = Query(...)):
    conn = get_db()
    conn.execute("DELETE FROM transactions WHERE id=? AND telegram_id=?", (tx_id, telegram_id))
    conn.commit()
    conn.close()
    return {"ok": True}

# ─── CATEGORIES ───────────────────────────────────────────
@app.get("/categories")
def get_categories(telegram_id: str):
    conn = get_db()
    ensure_user_categories(telegram_id, conn)
    rows = conn.execute("""
        SELECT * FROM categories WHERE telegram_id=? ORDER BY cat_type, position
    """, (telegram_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/categories")
def add_category(c: CategoryIn):
    import uuid
    conn = get_db()
    max_pos = conn.execute("""
        SELECT COALESCE(MAX(position),0)+1 FROM categories
        WHERE telegram_id=? AND cat_type=?
    """, (c.telegram_id, c.cat_type)).fetchone()[0]
    cat_id = str(uuid.uuid4())[:8]
    conn.execute("INSERT INTO categories(id,telegram_id,cat_type,emoji,name,position) VALUES(?,?,?,?,?,?)",
                 (cat_id, c.telegram_id, c.cat_type, c.emoji, c.name, max_pos))
    conn.commit()
    conn.close()
    return {"ok": True, "id": cat_id}

@app.delete("/categories/{cat_id}")
def delete_category(cat_id: str, telegram_id: str = Query(...)):
    conn = get_db()
    conn.execute("DELETE FROM categories WHERE id=? AND telegram_id=?", (cat_id, telegram_id))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.post("/categories/reorder")
def reorder_categories(r: CategoryReorder):
    conn = get_db()
    for i, cat_id in enumerate(r.ordered_ids):
        conn.execute("UPDATE categories SET position=? WHERE id=? AND telegram_id=?",
                     (i, cat_id, r.telegram_id))
    conn.commit()
    conn.close()
    return {"ok": True}

# ─── WISHES ───────────────────────────────────────────────
@app.get("/wishes")
def get_wishes(telegram_id: str):
    conn = get_db()
    rows = conn.execute("SELECT * FROM wishes WHERE telegram_id=? ORDER BY created_at", (telegram_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/wishes")
def add_wish(w: WishIn):
    conn = get_db()
    cur = conn.execute("INSERT INTO wishes(telegram_id,emoji,name,price) VALUES(?,?,?,?)",
                       (w.telegram_id, w.emoji, w.name, w.price))
    conn.commit()
    wish_id = cur.lastrowid
    conn.close()
    return {"ok": True, "id": wish_id}

@app.delete("/wishes/{wish_id}")
def delete_wish(wish_id: int, telegram_id: str = Query(...)):
    conn = get_db()
    conn.execute("DELETE FROM wishes WHERE id=? AND telegram_id=?", (wish_id, telegram_id))
    conn.commit()
    conn.close()
    return {"ok": True}

# ─── EXCEL EXPORT ─────────────────────────────────────────
@app.get("/export/excel")
def export_excel(telegram_id: str):
    conn = get_db()
    user = conn.execute("SELECT name FROM users WHERE telegram_id=?", (telegram_id,)).fetchone()
    user_name = user["name"] if user else "Пользователь"

    txs = conn.execute("""
        SELECT t.*, c.emoji, c.name as cat_name
        FROM transactions t
        LEFT JOIN categories c ON t.cat_id=c.id
        WHERE t.telegram_id=?
        ORDER BY t.date DESC
    """, (telegram_id,)).fetchall()

    wishes = conn.execute("SELECT * FROM wishes WHERE telegram_id=?", (telegram_id,)).fetchall()
    conn.close()

    wb = Workbook()

    # ── Sheet 1: Transactions ──
    ws = wb.active
    ws.title = "Операции"

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    green_font = Font(color="4ECB8D", bold=True)
    red_font = Font(color="F06B6B", bold=True)
    center = Alignment(horizontal="center")
    thin = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    ws.append([f"Pulim — {user_name}", "", "", "", ""])
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(bold=True, size=14, color="1A1A2E")
    ws["A1"].alignment = center
    ws.append([])

    headers = ["Дата", "Тип", "Название", "Категория", "Сумма (сум)"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    for tx in txs:
        row = [
            tx["date"],
            "Доход" if tx["type"] == "income" else "Расход",
            tx["name"],
            f"{tx['emoji'] or ''} {tx['cat_name'] or ''}".strip(),
            tx["amount"],
        ]
        ws.append(row)
        r = ws.max_row
        ws.cell(r, 5).number_format = '#,##0'
        ws.cell(r, 5).font = green_font if tx["type"] == "income" else red_font
        for col in range(1, 6):
            ws.cell(r, col).border = thin
            ws.cell(r, col).alignment = Alignment(horizontal="center" if col in [1,2,5] else "left")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # ── Sheet 2: Summary by month ──
    ws2 = wb.create_sheet("По месяцам")
    ws2.append(["Год", "Месяц", "Доходы (сум)", "Расходы (сум)", "Остаток (сум)"])
    for col in range(1, 6):
        ws2.cell(1, col).fill = header_fill
        ws2.cell(1, col).font = header_font
        ws2.cell(1, col).alignment = center

    months_data = {}
    for tx in txs:
        dt = tx["date"][:7]
        if dt not in months_data:
            months_data[dt] = {"income": 0, "expense": 0}
        if tx["type"] == "income":
            months_data[dt]["income"] += tx["amount"]
        else:
            months_data[dt]["expense"] += tx["amount"]

    for ym in sorted(months_data.keys(), reverse=True):
        y, m = ym.split("-")
        inc = months_data[ym]["income"]
        exp = months_data[ym]["expense"]
        ws2.append([int(y), int(m), inc, exp, inc - exp])
        r = ws2.max_row
        for col in range(1, 6):
            ws2.cell(r, col).number_format = '#,##0' if col > 2 else 'General'
            ws2.cell(r, col).alignment = center

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ── Sheet 3: Wishes ──
    ws3 = wb.create_sheet("Желания")
    ws3.append(["Эмодзи", "Название", "Стоимость (сум)"])
    for col in range(1, 4):
        ws3.cell(1, col).fill = header_fill
        ws3.cell(1, col).font = header_font
        ws3.cell(1, col).alignment = center
    for w in wishes:
        ws3.append([w["emoji"], w["name"], w["price"]])
        r = ws3.max_row
        ws3.cell(r, 3).number_format = '#,##0'
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 28
    ws3.column_dimensions["C"].width = 20

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"pulim_{telegram_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
